"""
Conversor de XLSX a CSV con limpieza automática de metadata.
"""
import openpyxl
import csv
import argparse
import os
from typing import Optional


def detect_header_row(sheet) -> int:
    """
    Detecta la fila donde comienzan los datos reales (después de metadata).
    
    Args:
        sheet: Hoja de Excel
        
    Returns:
        Número de fila donde comienza el header
    """
    for idx, row in enumerate(sheet.iter_rows(max_row=20), start=1):
        # Buscar fila con múltiples celdas no vacías (probable header)
        non_empty = sum(1 for cell in row if cell.value is not None)
        if non_empty >= 3:
            return idx
    return 1


def xlsx_to_csv(xlsx_path: str, csv_path: Optional[str] = None, auto_clean: bool = True) -> bool:
    """
    Convierte un archivo XLSX a CSV.
    
    Args:
        xlsx_path: Ruta del archivo XLSX
        csv_path: Ruta del archivo CSV de salida (opcional)
        auto_clean: Si True, detecta y elimina metadata automáticamente
        
    Returns:
        True si la conversión fue exitosa
    """
    if csv_path is None:
        csv_path = os.path.splitext(xlsx_path)[0] + '.csv'
    
    try:
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True)
        sheet = workbook.active
        
        start_row = detect_header_row(sheet) if auto_clean else 1
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if idx >= start_row:
                    writer.writerow(row)
        
        workbook.close()
        print(f"Conversión exitosa: {csv_path}")
        return True
        
    except Exception as e:
        print(f"Error al convertir {xlsx_path}: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(description='Convertir XLSX a CSV')
    parser.add_argument('input', help='Archivo XLSX de entrada')
    parser.add_argument('-o', '--output', help='Archivo CSV de salida')
    parser.add_argument('--no-clean', action='store_true', 
                       help='Desactivar limpieza automática de metadata')
    
    args = parser.parse_args()
    
    xlsx_to_csv(args.input, args.output, auto_clean=not args.no_clean)


if __name__ == "__main__":
    main()
